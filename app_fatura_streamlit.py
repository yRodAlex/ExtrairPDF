import streamlit as st
import pdfplumber
import pandas as pd
import re
import plotly.express as px
from io import BytesIO
from openpyxl import Workbook

st.set_page_config(page_title="Gestão Completa de Faturas e DRE", layout="wide")

st.title("💼 Faturas e Análises de DRE")

menu = st.sidebar.radio("Menu", ["📁 Converter Fatura PDF → DRE", "📊 Analisar DRE Consolidado"])


# ---------------- Função Agrupamento Avançado Itaú -----------------

def extrair_lancamentos_itau_avancado(pdf_path):
    datas, estabelecimentos, cidades, valores = [], [], [], []

    with pdfplumber.open(pdf_path) as pdf:
        for pagina in pdf.pages:
            palavras = pagina.extract_words()

            linhas_agrupadas = {}
            tolerancia_altura = 2  # Permite pequenas variações de altura

            for palavra in palavras:
                y_coord = round(palavra['top'] / tolerancia_altura) * tolerancia_altura
                if y_coord not in linhas_agrupadas:
                    linhas_agrupadas[y_coord] = []
                linhas_agrupadas[y_coord].append(palavra)

            for y in sorted(linhas_agrupadas.keys()):
                linha = sorted(linhas_agrupadas[y], key=lambda x: x['x0'])

                data = ""
                descricao = ""
                valor = ""

                for palavra in linha:
                    texto = palavra['text']
                    x0 = palavra['x0']

                    if re.match(r'\d{2}/\d{2}', texto) and x0 < 100:
                        data = texto
                    elif re.match(r'-?\d{1,3}(?:\.\d{3})*,\d{2}$', texto) and x0 > 400:
                        valor = texto
                    else:
                        descricao += texto + " "

                if data and valor and descricao.strip():
                    try:
                        valor_float = float(valor.replace('.', '').replace(',', '.'))
                    except:
                        continue

                    datas.append(data)
                    estabelecimentos.append(descricao.strip())
                    cidades.append("")
                    valores.append(valor_float)

    return datas, estabelecimentos, cidades, valores


# ---------------- Aba de Transformação PDF → DRE -----------------

if menu == "📁 Converter Fatura PDF → DRE":
    st.header("Conversão de Fatura para DRE (Excel)")

    banco = st.selectbox("Selecione o Banco:", ["itau", "sicoob"])
    mes = st.text_input("Mês da fatura (número ou nome, ex: 06 ou Junho):")
    ano = st.text_input("Ano da fatura (ex: 2025):")
    uploaded_file = st.file_uploader("Envie o PDF da fatura:", type=["pdf"])

    if uploaded_file and mes and ano:
        datas, estabelecimentos, cidades, valores = [], [], [], []

        if banco == "itau":
            caminho_temp = "temp_fatura.pdf"
            with open(caminho_temp, "wb") as f:
                f.write(uploaded_file.read())

            datas, estabelecimentos, cidades, valores = extrair_lancamentos_itau_avancado(caminho_temp)

        elif banco == "sicoob":
            with pdfplumber.open(uploaded_file) as pdf:
                lendo = False
                meses_dict = {'JAN':'01','FEV':'02','MAR':'03','ABR':'04','MAI':'05','JUN':'06',
                              'JUL':'07','AGO':'08','SET':'09','OUT':'10','NOV':'11','DEZ':'12'}

                for pagina in pdf.pages:
                    texto = pagina.extract_text()
                    if texto:
                        linhas = texto.split('\n')

                        for linha in linhas:
                            if "DATA" in linha and "DESCRIÇÃO" in linha and "VALOR" in linha:
                                lendo = True
                                continue

                            if lendo:
                                if "TOTAL" in linha:
                                    break

                                partes = linha.strip().split()
                                if len(partes) < 5:
                                    continue

                                dia = partes[0]
                                mes_abrev = partes[1].upper()
                                mes_num = meses_dict.get(mes_abrev, "00")
                                data_formatada = f"{dia}/{mes_num}"

                                valor_bruto = partes[-1].replace('.', '').replace(',', '.').replace('R$', '')
                                try:
                                    valor_float = float(valor_bruto)
                                except:
                                    continue

                                cidade = partes[-2].replace("R$", "").strip()
                                descricao = " ".join(partes[2:-2])

                                datas.append(data_formatada)
                                estabelecimentos.append(descricao.strip())
                                cidades.append(cidade.strip())
                                valores.append(valor_float)

        if datas:
            st.success(f"Lançamentos extraídos: {len(datas)}")

            df_resultado = pd.DataFrame({
                "Data": datas,
                "Estabelecimento": estabelecimentos,
                "Cidade": cidades,
                "Valor (R$)": valores
            })
            st.dataframe(df_resultado)

            output = BytesIO()
            wb = Workbook()
            nome_aba = f'{banco}-{mes}{ano}'
            ws = wb.active
            ws.title = nome_aba

            ws['A1'] = f'Fatura do mês {mes}, ano {ano}, Banco {banco.upper()}'
            ws.append([])
            ws.append([])
            ws.append(['Data', 'Estabelecimento', 'Cidade', 'Valor (R$)', 'Código Conta', 'Descrição Conta'])

            for i in range(len(datas)):
                linha_excel = i + 5
                ws.append([
                    datas[i],
                    estabelecimentos[i],
                    cidades[i],
                    valores[i],
                    '',
                    f'=VLOOKUP(D{linha_excel};\'Plano de Contas\'!A:B;2;FALSE)'
                ])

            ws2 = wb.create_sheet('Plano de Contas')
            ws2.append(['Código Conta', 'Descrição Conta'])
            ws2.append([1, 'Gasto Geral'])
            ws2.append([2, 'Restaurantes'])
            ws2.append([3, 'Transporte'])
            ws2.append([4, 'Mercado'])

            wb.save(output)
            output.seek(0)

            st.download_button(
                label="📥 Baixar Excel DRE Gerado",
                data=output,
                file_name=f'DRE_{banco}_{mes}_{ano}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            st.warning("Nenhum lançamento encontrado no PDF. Verifique o arquivo e tente novamente.")


# ---------------- Aba de Análise do DRE Consolidado -----------------

if menu == "📊 Analisar DRE Consolidado":
    st.header("Análise Exclusiva da aba 'DRE Consolidado'")

    arquivo = st.file_uploader("Importe o arquivo DRE (.xlsx ou .xlsm) com a aba 'DRE Consolidado':", type=["xlsx", "xlsm"])

    if arquivo:
        try:
            df = pd.read_excel(arquivo, sheet_name="DRE Consolidado")
            df.columns = df.columns.str.replace(r'R\$\s*', '', regex=True).str.strip()

            if "Descrição Conta" in df.columns:
                meses_colunas = [col for col in df.columns if re.match(r'.*/\d{2,4}', str(col))]

                if not meses_colunas:
                    st.warning("Não foram encontradas colunas de meses (ex: jun/25, jul/25).")
                else:
                    df = df.dropna(subset=["Descrição Conta"])
                    st.dataframe(df)

                    st.header("📊 Gastos por Categoria (Total por Mês)")
                    df_melt = df.melt(id_vars=["Descrição Conta"], value_vars=meses_colunas,
                                      var_name="Mês/Ano", value_name="Valor (R$)")

                    df_melt["Valor (R$)"] = df_melt["Valor (R$)"].replace({"R\\$": "", ",": "."}, regex=True).astype(float)

                    grafico = px.bar(df_melt, x="Descrição Conta", y="Valor (R$)", color="Mês/Ano",
                                     title="Comparativo de Gastos por Categoria e Mês", barmode="group")
                    st.plotly_chart(grafico)

                    st.header("🎯 Comparação Total por Categoria")
                    col_total = [col for col in df.columns if "Total" in col]
                    if col_total:
                        df_total = df[["Descrição Conta"] + col_total]
                        fig_pie = px.pie(df_total, names="Descrição Conta", values=col_total[0],
                                         title="Distribuição dos Gastos Totais por Categoria")
                        st.plotly_chart(fig_pie)

                    st.header("🔮 Previsão e Metas de Economia")
                    gasto_mensal = df_melt.groupby("Mês/Ano")["Valor (R$)"].sum().reset_index()
                    st.line_chart(gasto_mensal.set_index("Mês/Ano"))

                    media_gasto = gasto_mensal["Valor (R$)"].mean()
                    st.info(f"Gasto médio mensal atual: R$ {media_gasto:.2f}")

                    economia = st.number_input("Quanto pretende economizar por mês (R$):", min_value=0.0, step=50.0)
                    previsao_final_ano = 12 * (media_gasto - economia)

                    st.success(f"Se atingir essa economia, previsão de gasto anual: R$ {previsao_final_ano:.2f}")

            else:
                st.warning("A aba 'DRE Consolidado' não possui a coluna 'Descrição Conta'.")

        except Exception as e:
            st.error(f"Erro ao processar o arquivo: {e}")
